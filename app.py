"""
Activa IT - Descargador automático de cartas glosa (Previsora SOAT)
Versión mejorada para Railway: Detener/Reiniciar, Carpetas por IPS, Excel, Doble búsqueda
"""

import os
import re
import json
import time
import threading
import logging
from datetime import datetime
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_from_directory

# Para generar Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl no instalado. No se generará el archivo Excel.")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

BASE_DIR = Path(__file__).parent
DOWNLOAD_DIR = BASE_DIR / "downloads"
DOWNLOAD_DIR.mkdir(exist_ok=True)

# ==================== MAPA DE IPS POR NIT ====================
MAPA_IPS = {
    "900267064": "INVERSIONES_AZALUD_CLINICA_BAHIA",
    "900827065": "CENTRO_DE_DIAGNOSTICO_E_IMAGENES_BAHIA",
    "900657731": "CENTRO_MEDICO_Y_DE_REHABILITACION_BAHIA",
    "900826509": "RED_DE_URGENCIAS_DEL_MAGDALENA",
    "900513306": "FUNDACION_MARIA_REINA",
    "900600550": "INVERSIONES_MEDICAS_BARU",
    "900954800": "CENTRO_MEDICO_Y_DE_REHABILITACION_BARU",
    "900631361": "INVERSIONES_MEDICAS_VALLESALUD",
    "900257333": "ODONTOTRANS",
    "901081281": "URGETRAUMA",
    "900792417": "RED_DE_URGENCIAS_DE_LA_COSTA_PACIFICA",
    "901959993": "CLINICA_CORDIALIDAD",
}

# ==================== ESTADO GLOBAL ====================
job_state = {
    "running": False,
    "stopping": False,
    "logs": [],
    "stats": {"total": 0, "descargadas": 0, "errores": 0},
    "finished": False,
    "error": None,
    "errores_detalle": [],
    "descargas_exitosas": [],
}
job_lock = threading.Lock()
current_browser = None
current_context = None

def log(msg, level="info"):
    ts = datetime.now().strftime("%H:%M:%S")
    entry = {"ts": ts, "msg": msg, "level": level}
    with job_lock:
        job_state["logs"].append(entry)
    if level == "error":
        logger.error(msg)
    else:
        logger.info(msg)

def reset_state():
    with job_lock:
        job_state["running"] = False
        job_state["stopping"] = False
        job_state["logs"] = []
        job_state["stats"] = {"total": 0, "descargadas": 0, "errores": 0}
        job_state["finished"] = False
        job_state["error"] = None
        job_state["errores_detalle"] = []
        job_state["descargas_exitosas"] = []

def stop_job():
    global current_browser, current_context
    with job_lock:
        job_state["stopping"] = True
    if current_browser:
        try:
            current_browser.close()
            log("🛑 Navegador cerrado por solicitud de detención.")
        except:
            pass
    log("🛑 Proceso detenido por el usuario.")

# ==================== GENERADOR DE EXCEL ====================
def generar_reporte_excel(dl_dir, periodo, ips_nombre, exitosas, errores):
    if not EXCEL_AVAILABLE:
        return None
    excel_path = dl_dir / ips_nombre / f"reporte_{periodo}.xlsx"
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ws_exit = wb.active
    ws_exit.title = "Descargadas"
    ws_exit.append(["N° Factura", "Estado", "IPS", "Archivo Descargado", "Fecha/Hora"])
    for ex in exitosas:
        ws_exit.append([ex.get("factura"), ex.get("estado"), ips_nombre, ex.get("archivo"), ex.get("timestamp")])

    ws_err = wb.create_sheet("Errores")
    ws_err.append(["N° Factura", "Estado", "IPS", "Error", "Captura pantalla", "Fecha/Hora"])
    for err in errores:
        ws_err.append([err.get("factura"), err.get("estado"), ips_nombre, err.get("error"), err.get("captura"), err.get("timestamp")])

    wb.save(excel_path)
    return excel_path

# ==================== FUNCIONES AUXILIARES ====================

def _find_frame_with_text(page, regex_text: str):
    js = f"() => {{ const re = new RegExp({json.dumps(regex_text)}, 'i'); return re.test(document.body?.innerText || ''); }}"
    for fr in page.frames:
        try:
            if fr.evaluate(js):
                return fr
        except:
            continue
    return None

def _cerrar_traza_factura(page):
    js = """
        () => {
            const headers = document.querySelectorAll('.ui-dialog-titlebar, .modal-header, [class*="header"]');
            for (const h of headers) {
                if (h.textContent && h.textContent.includes('Traza de Factura')) {
                    const dlg = h.closest('.ui-dialog, .modal, [role="dialog"]');
                    if (dlg) {
                        const closeBtn = dlg.querySelector('.ui-dialog-titlebar-close, button.close, [aria-label*="lose"], [class*="close"]');
                        if (closeBtn) { closeBtn.click(); return true; }
                    }
                }
            }
            return false;
        }
    """
    for fr in page.frames:
        try:
            if fr.evaluate(js):
                time.sleep(0.5)
                return
        except:
            continue

def _extraer_nombre_ips(page, target_frame):
    def _buscar_en_frame(frame):
        try:
            js = """
                () => {
                    const body = document.body.innerText;
                    const match = body.match(/NIT\\s*:\\s*([\\d\\-\\s]+)/i);
                    if (match) {
                        let nit = match[1].replace(/[^0-9]/g, '');
                        return nit;
                    }
                    return "";
                }
            """
            nit = frame.evaluate(js).strip()
            if nit:
                return nit
        except:
            pass
        return ""

    nit = ""
    try:
        nit = _buscar_en_frame(page)
    except:
        pass
    if not nit:
        try:
            nit = _buscar_en_frame(target_frame)
        except:
            pass
    if not nit:
        for fr in page.frames:
            if fr != target_frame and fr != page:
                nit = _buscar_en_frame(fr)
                if nit:
                    break

    log(f"    🔍 NIT detectado: {nit}")

    if nit and nit in MAPA_IPS:
        nombre = MAPA_IPS[nit]
        log(f"    🏥 Nombre obtenido del mapa para NIT {nit}: {nombre}")
        return nombre

    js_nombre = """
        () => {
            const keywords = ["IPS","CLINICA","HOSPITAL","CENTRO","FUNDACIÓN","URGENCIAS","SALUD","ODONTOTRANS","URGETRAUMA","CORDIALIDAD"];
            const elementos = document.querySelectorAll('h1,h2,h3,h4,p,div');
            for (const el of elementos) {
                let txt = el.innerText.trim();
                if (txt.length > 5 && txt.length < 100) {
                    for (let kw of keywords) {
                        if (txt.toUpperCase().includes(kw)) {
                            return txt;
                        }
                    }
                }
            }
            return "";
        }
    """
    nombre = ""
    try:
        nombre = page.evaluate(js_nombre).strip()
    except:
        pass
    if not nombre:
        try:
            nombre = target_frame.evaluate(js_nombre).strip()
        except:
            pass

    if not nombre:
        nombre = "IPS_DESCONOCIDA"

    nombre = re.sub(r'[\\/*?:"<>|]', "", nombre)
    nombre = nombre.strip().replace(" ", "_")

    log(f"    🏥 IPS final: {nombre} (NIT: {nit})")
    return nombre


# ==================== FUNCIÓN _download_factura ====================
def _download_factura(page, context, modal_frame, fac: dict, dl_dir: Path, ips_nombre: str):
    import re
    num = fac["num"]
    tipo = fac["tipo"]
    target_label = "ActaDevolucion" if tipo == "devolucion" else "Envios_D"
    target_label_norm = target_label.replace('í', 'i')
    subcarpeta = "Auditada" if tipo == "auditada" else "Devolucion"

    ips_dir = dl_dir / ips_nombre
    dl_subdir = ips_dir / subcarpeta
    dl_subdir.mkdir(parents=True, exist_ok=True)

    bot_id = fac.get("botId")
    log(f"    🔗 Abriendo factura {num}...")
    num_solo_digitos = re.sub(r'\D', '', str(num))

    js_click_robusto = f"""
        () => {{
            const botId = '{bot_id}';
            const targetDigits = '{num_solo_digitos}';
            const fila = document.querySelector(`[data-bot-row-id="${{botId}}"]`);
            if (!fila) return {{ ok: false, reason: "fila_no_encontrada" }};
            fila.scrollIntoView({{block: 'center'}});
            function dispararClick(el) {{
                if (!el) return false;
                try {{ el.click(); }} catch (e) {{}}
                try {{ el.dispatchEvent(new MouseEvent('click', {{bubbles: true, cancelable: true, view: window}})); }} catch (e) {{}}
                return true;
            }}
            const candidatos = [];
            for (const a of fila.querySelectorAll('a')) {{
                const t = (a.textContent || '').trim();
                if (t.replace(/\\D/g, '') === targetDigits || candidatos.length === 0)
                    candidatos.push({{ tipo: 'a', el: a }});
            }}
            for (const el of fila.querySelectorAll('[onclick]')) {{
                if (!candidatos.find(c => c.el === el)) candidatos.push({{ tipo: 'onclick', el }});
            }}
            candidatos.push({{ tipo: 'fila', el: fila }});
            for (const td of fila.querySelectorAll('td')) candidatos.push({{ tipo: 'td', el: td }});
            for (const c of candidatos) dispararClick(c.el);
            return {{ ok: true, clickedWith: 'cascada', candidates: candidatos.length }};
        }}
    """
    result = None
    try:
        result = modal_frame.evaluate(js_click_robusto)
    except Exception as e:
        log(f"    ⚠️ Click falló: {e}", "warn")
    if not result or not result.get("ok"):
        for fr in page.frames:
            try:
                r = fr.evaluate(js_click_robusto)
                if r and r.get("ok"):
                    result = r
                    break
            except:
                continue
    if not result or not result.get("ok"):
        raise Exception(f"Click totalmente fallido para factura {num}.")
    log(f"    ✓ Click en factura {num} OK.")
    time.sleep(1.5)

    detalle_state = None
    detalle_frame = None
    for _ in range(60):
        if job_state.get("stopping"): return
        f = _find_frame_with_text(page, "Adjuntos por Factura")
        if f:
            try:
                has_traza = f.evaluate("() => /Traza de Factura/i.test(document.body?.innerText || '')")
                detalle_state = "traza" if has_traza else "adjuntos_directo"
            except:
                detalle_state = "adjuntos_directo"
            detalle_frame = f
            break
        f = _find_frame_with_text(page, "Traza de Factura")
        if f:
            detalle_state = "traza"
            detalle_frame = f
        time.sleep(0.5)
    if not detalle_frame:
        raise Exception("No apareció 'Traza de Factura' ni 'Adjuntos por Factura'.")
    time.sleep(1.5)
    log(f"    ✅ Detalle abierto (modo: {detalle_state}).")

    if detalle_state == "traza":
        log("    📑 Forzando cambio a pestaña 'Soportes'...")
        soportes_ok = False
        for intento in range(5):
            if job_state.get("stopping"): return
            for fr in page.frames:
                try:
                    has_tabs = fr.evaluate(r"""() => {
                        const txt = (document.body?.innerText || '').replace(/\n/g, ' ');
                        return /Factura.*Detalles.*Soportes/i.test(txt);
                    }""")
                    if has_tabs:
                        try:
                            fr.locator("text=Soportes").first.click(timeout=5000)
                            soportes_ok = True
                            break
                        except:
                            clicked = fr.evaluate("""() => {
                                for (const el of document.querySelectorAll('*')) {
                                    if ((el.textContent||'').trim() === 'Soportes') {
                                        el.click(); return true;
                                    }
                                }
                                return false;
                            }""")
                            if clicked:
                                soportes_ok = True
                                break
                except:
                    continue
            if soportes_ok:
                break
            time.sleep(1)
        if not soportes_ok:
            log("    ⚠️ No se pudo clickear Soportes", "warn")
        else:
            time.sleep(3)

    log("    ⏳ Esperando 'Adjuntos por Factura'...")
    adjuntos_frame = None
    for _ in range(90):
        if job_state.get("stopping"): return
        for fr in page.frames:
            try:
                if fr.evaluate("() => /Adjuntos por Factura|Buscar por.*Fecha/i.test(document.body?.innerText || '')"):
                    adjuntos_frame = fr
                    break
            except:
                continue
        if adjuntos_frame:
            break
        time.sleep(0.5)
    if not adjuntos_frame:
        raise Exception("No se encontró sección 'Adjuntos por Factura'.")
    for _ in range(35):
        if job_state.get("stopping"): return
        try:
            busy = adjuntos_frame.evaluate("() => /Procesando Solicitud/i.test(document.body?.innerText || '')")
            if not busy:
                break
        except:
            pass
        time.sleep(1)
    time.sleep(1)
    log("    ✅ Adjuntos cargados.")

    search_frame = adjuntos_frame

    def _escribir_buscador(texto):
        search_frame.evaluate(f"""
            () => {{
                const target = '{texto.replace('í', 'i')}';
                const inputs = document.querySelectorAll('input');
                let searchInput = null;
                for (const input of inputs) {{
                    const ph = (input.placeholder || '').toLowerCase();
                    if (ph.includes('buscar') || ph.includes('filtrar') || ph.includes('nombre')) {{
                        searchInput = input;
                        break;
                    }}
                }}
                if (!searchInput) return;
                searchInput.focus();
                searchInput.select();
                const nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                if (nativeSetter) nativeSetter.call(searchInput, target);
                else searchInput.value = target;
                searchInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
                searchInput.dispatchEvent(new Event('change', {{ bubbles: true }}));
                let parent = searchInput.closest('div, td, form, span');
                if (parent) {{
                    const btns = parent.querySelectorAll('button, a, [role="button"], span');
                    for (const btn of btns) {{
                        const html = (btn.outerHTML || '').toLowerCase();
                        const title = (btn.title || '').toLowerCase();
                        if (html.includes('search') || html.includes('lup') || title.includes('search')) {{
                            btn.click(); return;
                        }}
                    }}
                }}
            }}
        """)
        time.sleep(2)
        for _ in range(40):
            if job_state.get("stopping"): return
            processing = False
            for fr in page.frames:
                try:
                    if fr.evaluate("() => /Procesando Solicitud/i.test(document.body?.innerText || '')"):
                        processing = True
                        break
                except:
                    pass
            if not processing:
                break
            time.sleep(0.5)
        time.sleep(2)

    log(f"    🔍 Buscando '{target_label}'...")
    _escribir_buscador(target_label)
    archivo_seleccionado = False
    posibles_nombres = list({target_label, target_label_norm})

    for intento in range(4):
        if job_state.get("stopping"): return
        for fr in page.frames:
            try:
                resultado = fr.evaluate(f"""
                    () => {{
                        const nombres = {json.dumps(posibles_nombres)};
                        let contenedor = null;
                        const elementos = document.querySelectorAll('td, div, span, li, p, tr');
                        for (const el of elementos) {{
                            const txt = (el.innerText || '').trim();
                            for (const nombre of nombres) {{
                                if (txt === nombre) {{
                                    contenedor = el.closest('div[class*="file"], li[class*="file"], tr, div[class*="item"], div[class*="attach"], div[class*="row"]');
                                    if (!contenedor) contenedor = el.closest('div, li, tr');
                                    break;
                                }}
                            }}
                            if (contenedor) break;
                        }}
                        if (!contenedor) return {{ ok: false }};
                        let check = contenedor.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                        if (!check) check = contenedor.parentElement?.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                        if (check) {{
                            if (!check.checked) {{
                                check.click();
                                check.checked = true;
                                check.dispatchEvent(new Event('change', {{ bubbles: true }}));
                            }}
                            return {{ ok: true, metodo: 'checkbox' }};
                        }}
                        let iconoPdf = contenedor.querySelector('img[src*="pdf"], svg[aria-label*="pdf"], i[class*="pdf"], i[class*="file"], div[class*="pdf-icon"]');
                        if (iconoPdf) {{
                            iconoPdf.click();
                            return {{ ok: true, metodo: 'icono_pdf' }};
                        }}
                        contenedor.click();
                        contenedor.dispatchEvent(new MouseEvent('click', {{ bubbles: true, cancelable: true }}));
                        contenedor.dispatchEvent(new MouseEvent('dblclick', {{ bubbles: true, cancelable: true }}));
                        return {{ ok: true, metodo: 'contenedor_forzado' }};
                    }}
                """)
                if resultado and resultado.get('ok'):
                    log(f"    ✅ Selección realizada (método: {resultado.get('metodo')})")
                    archivo_seleccionado = True
                    break
            except Exception as e:
                log(f"    ⚠️ Error en intento {intento+1}: {e}", "warn")
        if archivo_seleccionado:
            break
        log(f"    🔄 Reintentando selección ({intento+1}/4)...")
        time.sleep(2)

    if not archivo_seleccionado:
        log(f"    ⚠️ No se encontró '{target_label}'. Intentando con 'Carta de Objeción'...")
        _escribir_buscador("Carta de Objeción")
        posibles_nombres_2 = ["Carta de Objeción", "Carta de Objeción".replace('ó', 'o')]
        for intento in range(4):
            if job_state.get("stopping"): return
            for fr in page.frames:
                try:
                    resultado = fr.evaluate(f"""
                        () => {{
                            const nombres = {json.dumps(posibles_nombres_2)};
                            let contenedor = null;
                            const elementos = document.querySelectorAll('td, div, span, li, p, tr');
                            for (const el of elementos) {{
                                const txt = (el.innerText || '').trim();
                                for (const nombre of nombres) {{
                                    if (txt === nombre) {{
                                        contenedor = el.closest('div[class*="file"], li[class*="file"], tr');
                                        if (!contenedor) contenedor = el.closest('div, li, tr');
                                        break;
                                    }}
                                }}
                                if (contenedor) break;
                            }}
                            if (!contenedor) return {{ ok: false }};
                            let check = contenedor.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                            if (!check) check = contenedor.parentElement?.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                            if (check) {{
                                if (!check.checked) {{
                                    check.click();
                                    check.checked = true;
                                    check.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                }}
                                return {{ ok: true, metodo: 'checkbox' }};
                            }}
                            contenedor.click();
                            contenedor.dispatchEvent(new MouseEvent('click', {{ bubbles: true, cancelable: true }}));
                            return {{ ok: true, metodo: 'contenedor_forzado' }};
                        }}
                    """)
                    if resultado and resultado.get('ok'):
                        log(f"    ✅ Selección realizada con 'Carta de Objeción'")
                        archivo_seleccionado = True
                        break
                except:
                    pass
            if archivo_seleccionado:
                break
            time.sleep(2)

    if not archivo_seleccionado:
        raise Exception(f"No se pudo seleccionar el archivo (intentó '{target_label}' y 'Carta de Objeción')")

    log("    ⏳ Esperando confirmación de selección...")
    for _ in range(20):
        if job_state.get("stopping"): return
        hay_error = False
        for fr in page.frames:
            try:
                if fr.evaluate("() => /Debe seleccionar por lo menos un documento/i.test(document.body?.innerText || '')"):
                    hay_error = True
                    break
            except:
                pass
        if not hay_error:
            log("    ✅ Selección confirmada")
            break
        time.sleep(1)

    log(f"    👁️ Buscando botón 'Abrir Documento'...")
    pdf_data = None
    pdf_url = None

    boton_encontrado = False
    start_time = time.time()
    while time.time() - start_time < 15:
        if job_state.get("stopping"): return
        for fr in page.frames:
            try:
                btn = fr.locator('button[title="Abrir Documento"], button[aria-label="Abrir Documento"], button:has(i.fa-eye), button:has(i.bi-eye)').first
                if btn.is_visible(timeout=2000):
                    boton_encontrado = True
                    break
            except:
                pass
        if boton_encontrado:
            break
        time.sleep(0.5)
    else:
        raise Exception("Botón 'Abrir Documento' no encontrado")

    for reintento in range(2):
        if job_state.get("stopping"): return
        new_page = None
        try:
            with context.expect_page(timeout=30000) as page_info:
                for fr in page.frames:
                    try:
                        btn = fr.locator('button[title="Abrir Documento"], button[aria-label="Abrir Documento"], button:has(i.fa-eye), button:has(i.bi-eye)').first
                        if btn.is_visible(timeout=5000):
                            for _ in range(10):
                                if btn.is_enabled():
                                    break
                                time.sleep(0.5)
                            btn.click()
                            log("    ✅ Clic en botón 'Abrir Documento'")
                            break
                    except:
                        pass
            new_page = page_info.value
            for _ in range(30):
                if job_state.get("stopping"): return
                url = new_page.url
                if url and url != "about:blank" and ("amazonaws" in url or ".pdf" in url.lower()):
                    pdf_url = url
                    break
                time.sleep(0.5)
        except Exception as e:
            log(f"    ⚠️ Intento {reintento+1}: No se abrió nueva pestaña: {e}", "warn")
        finally:
            if new_page:
                try:
                    new_page.close()
                except:
                    pass

        if pdf_url:
            try:
                response = context.request.get(pdf_url, timeout=60000)
                if response.ok:
                    pdf_data = response.body()
                    log(f"    ✅ PDF descargado ({len(pdf_data)//1024} KB)")
                    break
            except Exception as e:
                log(f"    ⚠️ Error descargando: {e}", "warn")

        if not pdf_data:
            log("    ⏳ Intentando descarga directa...")
            try:
                with page.expect_download(timeout=30000) as download_info:
                    for fr in page.frames:
                        try:
                            btn = fr.locator('button[title="Abrir Documento"], button:has(i.fa-eye), button:has(i.bi-eye)').first
                            if btn.is_visible(timeout=3000):
                                btn.click()
                                break
                        except:
                            pass
                download = download_info.value
                pdf_data = download.path().read_bytes() if download.path() else None
                log("    ✅ Descarga directa capturada")
                break
            except Exception as e:
                log(f"    ⚠️ No se capturó descarga: {e}", "warn")

        if not pdf_data:
            log(f"    🔄 Reintento {reintento+1}/2...")
            time.sleep(2)

    if not pdf_data:
        raise Exception("No se pudo obtener el PDF")

    safe_name = re.sub(r"[^\w\-_.]", "_", f"{num}_{target_label}.pdf")
    out_path = dl_subdir / safe_name
    out_path.write_bytes(pdf_data)
    log(f"    💾 PDF guardado: {out_path.name} ({len(pdf_data)//1024} KB)")

    with job_lock:
        job_state["descargas_exitosas"].append({
            "factura": num,
            "estado": fac["estado"],
            "archivo": str(out_path),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })

    _cerrar_traza_factura(page)
    time.sleep(0.8)


# ==================== AUTOMATIZACIÓN PRINCIPAL ====================

def run_automation(usuario: str, password: str, periodo: str, download_path: str):
    from playwright.sync_api import sync_playwright
    global current_browser, current_context

    dl_dir = Path(download_path)
    dl_dir.mkdir(parents=True, exist_ok=True)
    ips_nombre_actual = "IPS_SIN_NOMBRE"

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=False)
            context = browser.new_context(accept_downloads=True, viewport={"width": 1500, "height": 900})
            page = context.new_page()
            current_browser = browser
            current_context = context

            log("🔐 Iniciando sesión en Activa IT...")
            if job_state.get("stopping"): return
            page.goto("https://activa-it.net/Login.aspx", wait_until="networkidle", timeout=60000)
            log(f"  → Usuario: {usuario}")
            page.fill('input[placeholder="Usuario"]', usuario)
            page.fill('input[placeholder="Contraseña"]', password)
            try:
                checkbox = page.locator('input[type="checkbox"]').first
                if not checkbox.is_checked():
                    checkbox.check()
            except:
                pass
            page.click('button:has-text("Inicio de sesión"), input[value="Inicio de sesión"]')
            page.wait_for_url("**/Index.aspx", timeout=60000)
            time.sleep(2)
            log("✅ Sesión iniciada correctamente.")
            if job_state.get("stopping"): return

            log("📂 Navegando a módulo BI IPS...")
            time.sleep(3)

            def _find_periodo_in_frames():
                js_check = f"""
                    () => {{
                        const bodyText = (document.body?.innerText || '').toLowerCase();
                        const periodo = '{periodo}'.toLowerCase();
                        if (bodyText.includes(periodo)) return true;
                        const variaciones = ['abr26', 'abr-26', 'abr.26', 'abr/26', 'abr2026'];
                        return variaciones.some(v => bodyText.includes(v));
                    }}
                """
                for fr in page.frames:
                    try:
                        if fr.evaluate(js_check):
                            return fr
                    except:
                        continue
                return None

            if job_state.get("stopping"): return
            clicked = False
            for intento in range(3):
                try:
                    page.locator("text=BI IPS").first.click(timeout=15000)
                    clicked = True
                    log("  ✓ Click directo en 'BI IPS' OK.")
                    break
                except:
                    pass
                try:
                    page.click("text=Inteligencia de Negocio", timeout=8000)
                    time.sleep(1)
                    page.click("text=BI IPS", timeout=8000)
                    clicked = True
                    log("  ✓ Click vía 'Inteligencia de Negocio' + 'BI IPS' OK.")
                    break
                except:
                    pass
                try:
                    page.click("[class*='menu-toggle'], [class*='hamburger'], .sidebar-toggle", timeout=5000)
                    time.sleep(2)
                    page.click("text=BI IPS", timeout=8000)
                    clicked = True
                    log("  ✓ Click vía hamburguesa + 'BI IPS' OK.")
                    break
                except Exception as e:
                    log(f"    ⚠️ Intento {intento+1} falló: {e}", "warn")
                    time.sleep(2)
            if not clicked:
                raise Exception("No se encontró el módulo BI IPS en el menú.")

            time.sleep(3)
            log("✅ Módulo BI IPS abierto. Buscando período...")
            target_frame = None
            for i in range(120):
                if job_state.get("stopping"): return
                target_frame = _find_periodo_in_frames()
                if target_frame:
                    log(f"✅ Período '{periodo}' detectado tras {(i+1)*0.5:.1f}s.")
                    break
                time.sleep(0.5)
            if not target_frame:
                raise Exception(f"No se pudo localizar el período '{periodo}' tras 60s.")

            log("🏥 Obteniendo nombre de la IPS...")
            ips_nombre_actual = _extraer_nombre_ips(page, target_frame)

            if job_state.get("stopping"): return
            log(f"📅 Click en columna Cant del período '{periodo}'...")
            click_result = target_frame.evaluate(f"""
                () => {{
                    const rows = document.querySelectorAll('tr');
                    for (const row of rows) {{
                        const cells = row.querySelectorAll('td');
                        if (cells.length < 3) continue;
                        const firstText = cells[0].textContent.trim();
                        if (firstText !== '{periodo}') continue;
                        const links = row.querySelectorAll('a');
                        if (links.length === 0) return {{ ok: false, reason: 'sin_links' }};
                        const firstLink = links[0];
                        const value = firstLink.textContent.trim();
                        if (value === '0') return {{ ok: false, reason: 'cant_cero', value: '0' }};
                        firstLink.scrollIntoView({{block: 'center'}});
                        firstLink.click();
                        return {{ ok: true, value: value }};
                    }}
                    return {{ ok: false, reason: 'fila_no_encontrada' }};
                }}
            """)
            if click_result.get("reason") == "cant_cero":
                log(f"ℹ️ El período '{periodo}' tiene 0 facturas radicadas.", "warn")
                browser.close()
                return
            if not click_result.get("ok"):
                raise Exception(f"No se pudo hacer click en Cant de '{periodo}': {click_result.get('reason')}")
            log(f"  → Click en Cant: {click_result.get('value')}")

            log("⏳ Esperando modal 'Listado de facturas recibidas'...")
            modal_frame = None
            for _ in range(60):
                if job_state.get("stopping"): return
                for fr in page.frames:
                    try:
                        if fr.evaluate("() => /Listado de facturas recibidas/i.test(document.body?.innerText || '')"):
                            modal_frame = fr
                            break
                    except:
                        continue
                if modal_frame:
                    break
                time.sleep(0.5)
            if not modal_frame:
                raise Exception("El modal 'Listado de facturas recibidas' no apareció.")

            log("⏳ Esperando datos del listado...")
            data_frame = None
            tiempo_espera = 0
            while tiempo_espera < 60:
                if job_state.get("stopping"): return
                for fr in page.frames:
                    try:
                        if fr.evaluate("() => /Pendiente de recibir Informaci|Devoluci[oó]n de entrada/i.test(document.body?.innerText || '')"):
                            data_frame = fr
                            break
                    except:
                        continue
                if data_frame:
                    break
                time.sleep(0.5)
                tiempo_espera += 0.5

            if not data_frame:
                log("⚠️ No se encontraron facturas con los estados objetivo.", "warn")
                browser.close()
                return

            log(f"✅ Datos detectados en frame '{data_frame.name or '(main)'}'.")
            time.sleep(2)

            log("🔍 Extrayendo facturas...")
            js_extract = r"""
            (state) => {
                const ESTADOS = [
                    { nombre: 'Auditada: Pendiente de recibir Informacion', regex: /auditada\s*:\s*pendiente\s+de\s+recibir\s+informaci[oó]n/i, tipo: 'auditada' },
                    { nombre: 'En radicacion: Devolución de entrada', regex: /en\s+radicaci[oó]n\s*:\s*devoluci[oó]n\s+de\s+entrada/i, tipo: 'devolucion' },
                    { nombre: 'En auditoria: Pendiente de informar Orden de pago al Pagador', regex: /en\s+auditori?a\s*:\s*pendiente\s+de\s+informar\s+orden\s+de\s+pago\s+al\s+pagador/i, tipo: 'auditada' },
                ];
                const filas = document.querySelectorAll('tr, [role="row"], li');
                const nuevas = [];
                for (const fila of filas) {
                    const fullText = (fila.innerText || '').replace(/\s+/g, ' ').trim();
                    if (!fullText || fullText.length < 20 || fullText.length > 400) continue;
                    if (!/\d{2}\/\d{2}\/\d{4}/.test(fullText)) continue;
                    let tipoDetectado = null, nombreEstado = null;
                    for (const e of ESTADOS) {
                        if (e.regex.test(fullText)) { tipoDetectado = e.tipo; nombreEstado = e.nombre; break; }
                    }
                    if (!tipoDetectado) continue;
                    const tokens = fullText.split(/\s+/);
                    const candidatosNum = tokens.filter(t => { const digits = t.replace(/\D/g, ''); return digits.length >= 6 && digits.length <= 10; });
                    if (candidatosNum.length === 0 || candidatosNum.length > 6) continue;
                    const numNorm = candidatosNum[0].replace(/\D/g, '');
                    if (state.seen.includes(numNorm)) continue;
                    const botId = 'bot_' + state.nextId;
                    state.nextId++;
                    fila.setAttribute('data-bot-row-id', botId);
                    nuevas.push({
                        botId: botId, num: numNorm, rawNum: candidatosNum[0],
                        tipo: tipoDetectado, estado: nombreEstado,
                        textoFila: fullText.slice(0, 150), tagName: fila.tagName.toLowerCase(),
                    });
                    state.seen.push(numNorm);
                }
                return { nuevas: nuevas, total: state.seen.length };
            }
            """
            extract_state = {"nextId": 0, "seen": []}
            facturas_acumuladas = []
            rondas_sin_nuevos = 0
            for ronda in range(20):
                if job_state.get("stopping"): return
                try:
                    res = data_frame.evaluate(js_extract, extract_state)
                except:
                    res = {"nuevas": []}
                nuevas = res.get("nuevas", [])
                if nuevas:
                    facturas_acumuladas.extend(nuevas)
                    rondas_sin_nuevos = 0
                    log(f"  Ronda {ronda+1}: +{len(nuevas)} (Total: {len(facturas_acumuladas)})")
                else:
                    rondas_sin_nuevos += 1
                extract_state["seen"] = list(set(extract_state["seen"] + [n["num"] for n in nuevas]))
                if rondas_sin_nuevos >= 5:
                    break
                try:
                    data_frame.evaluate("() => { const scrollables = document.querySelectorAll('div, table, tbody, [class*=\"scroll\"]'); for (const s of scrollables) { if (s.scrollHeight > s.clientHeight + 20) s.scrollTop += s.clientHeight * 0.8; } window.scrollBy(0, window.innerHeight * 0.8); }")
                except:
                    pass
                time.sleep(0.5)
            log(f"📊 {len(facturas_acumuladas)} facturas detectadas.")
            facturas_objetivo = facturas_acumuladas

            with job_lock:
                job_state["stats"]["total"] = len(facturas_objetivo)
            cnt_aud = sum(1 for f in facturas_objetivo if f["tipo"] == "auditada")
            cnt_dev = sum(1 for f in facturas_objetivo if f["tipo"] == "devolucion")
            log("📋 RESUMEN:")
            log(f"  • Auditada: {cnt_aud}")
            log(f"  • Devolucion: {cnt_dev}")
            log(f"  TOTAL: {len(facturas_objetivo)}")
            if not facturas_objetivo:
                log("ℹ️ No hay facturas con los estados requeridos.")
                browser.close()
                return

            for idx, fac in enumerate(facturas_objetivo, 1):
                if job_state.get("stopping"): return
                log(f"[{idx}/{len(facturas_objetivo)}] Factura {fac['num']} ({fac['tipo']})...")
                try:
                    _download_factura(page, context, data_frame, fac, dl_dir, ips_nombre_actual)
                    with job_lock:
                        job_state["stats"]["descargadas"] += 1
                    log(f"  ✅ Descargada: {fac['num']}", "success")
                except Exception as e:
                    with job_lock:
                        job_state["stats"]["errores"] += 1
                        error_msg = str(e)
                        if "No se pudo seleccionar el archivo" in error_msg:
                            error_msg = f"En la factura {fac['num']} no se encontró soporte {('Envios_D' if fac['tipo']=='auditada' else 'ActaDevolucion')}"
                        error_info = {
                            "factura": fac['num'],
                            "estado": fac['estado'],
                            "error": error_msg,
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        }
                        try:
                            errores_dir = dl_dir / ips_nombre_actual / "Errores"
                            errores_dir.mkdir(parents=True, exist_ok=True)
                            cap_path = errores_dir / f"ERROR_{fac['num']}.png"
                            page.screenshot(path=str(cap_path))
                            error_info["captura"] = str(cap_path)
                        except:
                            error_info["captura"] = ""
                        job_state["errores_detalle"].append(error_info)
                    log(f"  ⚠️ Error: {error_msg}", "error")
                    _cerrar_traza_factura(page)
                    time.sleep(1)

            browser.close()

            if job_state["errores_detalle"]:
                errores_dir = dl_dir / ips_nombre_actual / "Errores"
                errores_dir.mkdir(parents=True, exist_ok=True)
                with open(errores_dir / "_errores.txt", "w", encoding="utf-8") as f:
                    f.write("=== FACTURAS CON ERRORES ===\n")
                    for err in job_state["errores_detalle"]:
                        f.write(f"{err['timestamp']} - Factura {err['factura']}: {err['error']}\n")

            excel_path = generar_reporte_excel(dl_dir, periodo, ips_nombre_actual,
                                              job_state["descargas_exitosas"],
                                              job_state["errores_detalle"])
            if excel_path:
                log(f"📊 Reporte Excel generado: {excel_path}")

            log("🎉 Proceso completado.")

    except Exception as e:
        if not job_state.get("stopping"):
            log(f"💥 Error crítico: {e}", "error")
            with job_lock:
                job_state["error"] = str(e)
        else:
            log("Proceso detenido por el usuario.")
    finally:
        with job_lock:
            job_state["running"] = False
            job_state["finished"] = True
            job_state["stopping"] = False
        current_browser = None
        current_context = None


# ==================== RUTAS FLASK ====================

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/start", methods=["POST"])
def start_job():
    data = request.json or {}
    usuario = data.get("usuario", "").strip()
    password = data.get("password", "").strip()
    periodo = data.get("periodo", "").strip()
    custom_path = data.get("download_path", "").strip()
    if not all([usuario, password, periodo]):
        return jsonify({"ok": False, "error": "Faltan campos requeridos"}), 400
    with job_lock:
        if job_state["running"]:
            return jsonify({"ok": False, "error": "Ya hay un proceso en ejecución"}), 409
        job_state["running"] = True
        job_state["finished"] = False
        job_state["error"] = None
        job_state["stats"] = {"total": 0, "descargadas": 0, "errores": 0}
        job_state["errores_detalle"] = []
        job_state["descargas_exitosas"] = []
    dl_path = custom_path if custom_path else str(DOWNLOAD_DIR / periodo)
    t = threading.Thread(target=run_automation, args=(usuario, password, periodo, dl_path), daemon=True)
    t.start()
    return jsonify({"ok": True, "download_path": dl_path})

@app.route("/api/stop", methods=["POST"])
def stop_job_route():
    with job_lock:
        if not job_state["running"]:
            return jsonify({"ok": False, "message": "No hay proceso en ejecución"}), 400
    stop_job()
    return jsonify({"ok": True, "message": "Deteniendo proceso..."})

@app.route("/api/reset", methods=["POST"])
def reset_job_route():
    with job_lock:
        if job_state["running"]:
            stop_job()
            time.sleep(2)
    reset_state()
    return jsonify({"ok": True, "message": "Estado reiniciado."})

@app.route("/api/status")
def get_status():
    with job_lock:
        return jsonify({
            "running": job_state["running"],
            "finished": job_state["finished"],
            "error": job_state["error"],
            "stats": job_state["stats"],
            "logs": job_state["logs"][-200:],
        })

@app.route("/api/logs")
def get_logs():
    since = int(request.args.get("since", 0))
    with job_lock:
        return jsonify({"logs": job_state["logs"][since:]})

@app.route("/api/logs", methods=["DELETE"])
def clear_logs():
    with job_lock:
        job_state["logs"] = []
    return jsonify({"ok": True})

@app.route("/api/files")
def list_files():
    periodo = request.args.get("periodo", "")
    folder = DOWNLOAD_DIR / periodo if periodo else DOWNLOAD_DIR
    files = []
    if folder.exists():
        for f in sorted(folder.iterdir()):
            if f.is_file():
                files.append({"name": f.name, "size": f.stat().st_size, "path": str(f), "periodo": periodo})
    return jsonify({"files": files})

@app.route("/downloads/<path:filename>")
def download_file(filename):
    return send_from_directory(DOWNLOAD_DIR, filename, as_attachment=True)

@app.route("/api/periodos")
def get_periodos():
    periodos = []
    for d in DOWNLOAD_DIR.iterdir():
        if d.is_dir():
            count = len(list(d.glob("**/*.pdf")))
            periodos.append({"name": d.name, "count": count})
    return jsonify({"periodos": sorted(periodos, key=lambda x: x["name"], reverse=True)})

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    print("\n" + "=" * 55)
    print("  🏥 Activa IT — Descargador de Cartas Glosa (Railway)")
    print("=" * 55)
    print(f"  📂 Carpeta de descargas: {DOWNLOAD_DIR}")
    print(f"  🌐 Puerto: {port}")
    print("=" * 55 + "\n")
    app.run(host="0.0.0.0", port=port, debug=False)